import os
import random
import datetime
import openpyxl
from http import HTTPStatus
from dashscope import Generation  # 建议dashscope SDK 的版本 >= 1.14.0

# 无线网LTE基站退服的重大故障定义是什么？

# 介绍无线网BSC重大故障定义。

# 一个本地网（地市）内65个LTE基站同时阻断3小时，是否触发重大故障？
prompt1 = '''
#01 你是一个问答对数据集处理专家。

#02 你的任务是根据我给出的内容，生成适合作为问答对数据集的问题。

#03 问题要尽量短，不要太长。

#04 一句话中只能有一个问题。

#05 尽可能多生成适合作为问答对的问题。

#06 生成问题示例：

"""

无线网LTE基站退服的重大故障定义是什么？

介绍无线网BSC重大故障定义。

一个本地网（地市）内65个LTE基站同时阻断3小时，是否触发重大故障？

家客业务的重大故障定义是什么？

家客OLT设备同时退服重大故障定义。

同一地市10000个用户（含）以上的家客业务同时阻断超过2小时，是否触发重大故障？

"""

#07 以下是我给出的内容：

"""

{{此处替换成你的内容}}

"""
'''

prompt2 = '''
#01 你是一个问答对数据集处理专家。

#02 你的任务是根据我的问题和我给出的内容，生成对应的问答对。

#03 答案要全面，多使用我的信息，内容要更丰富。

#04 你必须根据我的问答对示例格式来生成：

"""

"问题": "家客业务的重大故障定义是什么？"|"答复": "同一地市10000个用户（含）以上的家客业务同时阻断超过1小时。"

"问题": "家客OLT设备同时退服重大故障定义。"|"答复": "同一地市超过5个(含)家客OLT设备同时退服，且退服时长超过2小时（备注：统计时段不包括1-6点）。"

"问题": "同一地市10000个用户（含）以上的家客业务同时阻断超过2小时，是否触发重大故障？"|"答复": "同一地市10000个用户（含）以上的家客业务同时阻断超过1小时。所以触发了重大故障。"

"问题": "BSC的重大故障定义是什么"|"答复": "BSC重大故障定义如下：任一BSC发生阻断超过60分钟。"

"问题": "2G基站的重大故障定义是什么"|"答复": "一个本地网（地市）内40个（含）以上2G基站（且载频总数超过500个（含））同时阻断3小时，或120个（含）以上2G基站同时阻断超过1小时。"

"问题": "一个本地网（地市）内65个LTE基站同时阻断3小时，是否触发重大故障？"|"答复": "一个本地网（地市）内60个（含）以上LTE基站同时阻断（含脱管类基站）3小时，或150个（含）以上LTE基站同时阻断（含脱管类基站）超过1小时，触发重大故障。"

"""

#05 我的问题如下：

"""

{{此处替换成你上一步生成的问题}}

"""

#06 我的内容如下：

"""

{{此处替换成你的内容}}

"""
'''


def generate_question(text_content):
    prompt = prompt1.replace("{{此处替换成你的内容}}", text_content)
    messages = [{'role': 'system', 'content': 'Your primary goal is to provide accurate and concise information.'},
                {'role': 'user', 'content': prompt}]
    response = Generation.call(model="qwen-turbo",
                               messages=messages,
                               # 设置随机数种子seed，如果没有设置，则随机数种子默认为1234
                               seed=random.randint(1, 10000),
                               # 将输出设置为"message"格式
                               result_format='message')
    if response.status_code == HTTPStatus.OK:
        # print(response['output'])
        # print(response['output']['choices'][0]['message']['content'])
        return response['output']['choices'][0]['message']['content']
    else:
        print('Request id: %s, Status code: %s, error code: %s, error message: %s' % (
            response.request_id, response.status_code,
            response.code, response.message
        ))


def generate_qa(text_content, question_text):
    prompt = prompt2.replace("{{此处替换成你上一步生成的问题}}", question_text).replace("{{此处替换成你的内容}}",
                                                                                        text_content)
    messages = [{'role': 'system', 'content': 'Your primary goal is to provide accurate and concise information.'},
                {'role': 'user', 'content': prompt}]
    response = Generation.call(model="qwen-turbo",
                               messages=messages,
                               # 设置随机数种子seed，如果没有设置，则随机数种子默认为1234
                               seed=random.randint(1, 10000),
                               # 将输出设置为"message"格式
                               result_format='message')
    if response.status_code == HTTPStatus.OK:
        # print(response['output'])
        # print(response['output']['choices'][0]['message']['content'])
        return response['output']['choices'][0]['message']['content']
        # return response['output']['choices']
    else:
        print('Request id: %s, Status code: %s, error code: %s, error message: %s' % (
            response.request_id, response.status_code,
            response.code, response.message
        ))



def write_to_file(content):
    timestamp = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
    file_name = f"./new_file/new_file_{timestamp}.txt"
    with open(file_name, "w", encoding='utf-8') as file:
        file.write(content)
    print("File 'new_file.txt' has been created and written.")
    with open(file_name, "r", encoding='utf-8') as file:
        content = file.readlines()
    return content


def read_file(file_name):
    try:
        with open(file_name, "r", encoding='utf-8') as file:
            content = file.read()
        return content
    except FileNotFoundError:
        print(f"File '{file_name}' not found.")


def write_to_excel(content, sheet_name):
    file_name = '政企业务.xlsx'
    if os.path.exists(file_name):
        book = openpyxl.load_workbook(file_name)
    else:
        book = openpyxl.Workbook()
    sheet = book.create_sheet(title=sheet_name)
    # sheet.append(['问题', '答复'])
    for ct in content:
        if ct == '\n':
            continue
        else:
            sheet.append([ct])
    book.save(file_name)


def main(i):
    text_content = read_file("input_file_政企业务.txt")
    print('text_content\n', text_content)
    question_text = generate_question(text_content=text_content)
    print('question_text\n', question_text)
    qa_text = generate_qa(text_content=text_content, question_text=question_text)
    print('qa_text\n', qa_text)
    content = write_to_file(qa_text)
    write_to_excel(content=content, sheet_name=i)


for i in range(1, 11):
    main(str(i))
